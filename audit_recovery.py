"""Audit, versioning, integrity validation, and recovery services."""

import hashlib
import io
import subprocess
import traceback
import uuid
from datetime import datetime

from bson.binary import Binary
from openpyxl import load_workbook
from pymongo import ASCENDING, DESCENDING
from pymongo.errors import DuplicateKeyError


AUDIT_COLLECTION = "audit_logs"
HISTORY_COLLECTION = "processing_history"
COMMERCIAL_AUDIT_COLLECTION = "commercial_audit"
VERSION_COLLECTION = "report_versions"
VALIDATION_COLLECTION = "validation_results"
FAILURE_COLLECTION = "failure_diagnostics"


def utc_now_iso():
    return datetime.utcnow().isoformat()


def workbook_checksum(xlsx_bytes):
    return hashlib.sha256(bytes(xlsx_bytes)).hexdigest()


def workbook_metadata(xlsx_bytes):
    workbook = load_workbook(io.BytesIO(bytes(xlsx_bytes)), read_only=True, data_only=True)
    try:
        sheet_rows = {
            worksheet.title: int(worksheet.max_row or 0)
            for worksheet in workbook.worksheets
        }
        return {
            "workbook_size_bytes": len(xlsx_bytes),
            "sheet_count": len(workbook.sheetnames),
            "rows_processed": sum(sheet_rows.values()),
            "sheet_rows": sheet_rows,
            "checksum_sha256": workbook_checksum(xlsx_bytes),
        }
    finally:
        workbook.close()


def ensure_audit_indexes(database):
    database[AUDIT_COLLECTION].create_index(
        [("timestamp", DESCENDING), ("event_id", ASCENDING)],
        name="idx_audit_recent",
    )
    database[AUDIT_COLLECTION].create_index(
        [("job_id", ASCENDING), ("timestamp", DESCENDING)],
        name="idx_audit_job",
    )
    database[AUDIT_COLLECTION].create_index(
        [("channel", ASCENDING), ("date", ASCENDING), ("timestamp", DESCENDING)],
        name="idx_audit_report",
    )
    database[HISTORY_COLLECTION].create_index(
        [("job_id", ASCENDING)],
        name="idx_history_job",
        unique=True,
        partialFilterExpression={"job_id": {"$type": "string"}},
    )
    database[HISTORY_COLLECTION].create_index(
        [("original_filename", ASCENDING), ("upload_time", DESCENDING)],
        name="idx_history_search",
    )
    database[HISTORY_COLLECTION].create_index(
        [("job_status", ASCENDING), ("upload_time", DESCENDING)],
        name="idx_history_status",
    )
    database[COMMERCIAL_AUDIT_COLLECTION].create_index(
        [("channel", ASCENDING), ("date", ASCENDING), ("timestamp", DESCENDING)],
        name="idx_commercial_audit_report",
    )
    database[VERSION_COLLECTION].create_index(
        [("channel", ASCENDING), ("date", ASCENDING), ("version_number", ASCENDING)],
        name="idx_report_versions",
        unique=True,
    )
    database[VALIDATION_COLLECTION].create_index(
        [("status", ASCENDING), ("timestamp", DESCENDING)],
        name="idx_validation_status",
    )
    database[VALIDATION_COLLECTION].create_index(
        [("channel", ASCENDING), ("date", ASCENDING), ("timestamp", DESCENDING)],
        name="idx_validation_report",
    )
    database[FAILURE_COLLECTION].create_index(
        [("timestamp", DESCENDING), ("error_category", ASCENDING)],
        name="idx_failures_recent",
    )
    database[FAILURE_COLLECTION].create_index(
        [("job_id", ASCENDING), ("timestamp", DESCENDING)],
        name="idx_failures_job",
    )


def record_audit(
    database,
    user_action,
    status,
    channel="",
    date="",
    job_id="",
    details=None,
    event_id=None,
):
    document = {
        "event_id": event_id or str(uuid.uuid4()),
        "timestamp": utc_now_iso(),
        "user_action": str(user_action),
        "channel": str(channel or ""),
        "date": str(date or ""),
        "job_id": str(job_id or ""),
        "status": str(status),
        "details": details or {},
    }
    database[AUDIT_COLLECTION].insert_one(document)
    return document


def classify_error(exc):
    name = type(exc).__name__
    if isinstance(exc, subprocess.TimeoutExpired) or "timeout" in name.lower():
        return "TIMEOUT_EXCEPTION"
    if "mongo" in name.lower() or "pymongo" in type(exc).__module__.lower():
        return "MONGODB_EXCEPTION"
    if "openpyxl" in type(exc).__module__.lower() or "workbook" in str(exc).lower():
        return "WORKBOOK_EXCEPTION"
    if "validation" in name.lower():
        return "VALIDATION_FAILURE"
    return "PYTHON_EXCEPTION"


def record_failure(database, exc, job_id="", context=None, category=None):
    document = {
        "failure_id": str(uuid.uuid4()),
        "timestamp": utc_now_iso(),
        "job_id": str(job_id or ""),
        "error_category": category or classify_error(exc),
        "exception_type": type(exc).__name__,
        "message": str(exc)[:4000],
        "stack_trace": traceback.format_exc()[-20000:],
        "context": context or {},
    }
    database[FAILURE_COLLECTION].insert_one(document)
    return document


def record_processing_history(database, job, **updates):
    now = utc_now_iso()
    upload_time = job.get("created_at") or now
    started_at = updates.get("started_at", job.get("started_at"))
    completed_at = updates.get("completed_at", job.get("completed_at"))
    duration_ms = updates.get("processing_duration_ms")
    if duration_ms is None and started_at and completed_at:
        try:
            duration_ms = round(
                (
                    datetime.fromisoformat(completed_at)
                    - datetime.fromisoformat(started_at)
                ).total_seconds()
                * 1000,
                1,
            )
        except (TypeError, ValueError):
            duration_ms = None

    document = {
        "job_id": str(job.get("job_id") or ""),
        "original_filename": job.get("original_filename") or "",
        "upload_time": upload_time,
        "processing_duration_ms": duration_ms,
        "workbook_size_bytes": updates.get(
            "workbook_size_bytes",
            job.get("result_size_bytes") or 0,
        ),
        "sheet_count": updates.get("sheet_count", 0),
        "rows_processed": updates.get("rows_processed", 0),
        "job_status": updates.get("job_status", job.get("status") or ""),
        "error_details": updates.get("error_details", job.get("error_message")),
        "channel": updates.get("channel", job.get("channel") or ""),
        "date": updates.get("date", job.get("date") or ""),
        "report_file_id": updates.get("report_file_id", job.get("report_file_id")),
        "updated_at": now,
    }
    database[HISTORY_COLLECTION].update_one(
        {"job_id": document["job_id"]},
        {
            "$set": document,
            "$setOnInsert": {"history_id": str(uuid.uuid4())},
        },
        upsert=True,
    )
    return document


def record_commercial_change(
    database,
    channel,
    date,
    operation_type,
    modified_field,
    before_state,
    after_state,
    user_action,
    job_id="",
):
    document = {
        "change_id": str(uuid.uuid4()),
        "timestamp": utc_now_iso(),
        "channel": channel,
        "date": date,
        "job_id": str(job_id or ""),
        "operation_type": operation_type,
        "modified_field": modified_field,
        "before_state": before_state,
        "after_state": after_state,
        "user_action": user_action,
    }
    database[COMMERCIAL_AUDIT_COLLECTION].insert_one(document)
    return document


def _next_version_number(database, channel, date):
    latest = database[VERSION_COLLECTION].find_one(
        {"channel": channel, "date": date},
        {"version_number": 1},
        sort=[("version_number", DESCENDING)],
    )
    return int(latest.get("version_number") or 0) + 1 if latest else 1


def create_report_version(
    database,
    channel,
    date,
    xlsx_bytes,
    original_filename,
    file_id,
    reason,
    modifications=None,
    job_id="",
    genre="",
    source_version=None,
):
    metadata = workbook_metadata(xlsx_bytes)
    for attempt in range(5):
        version_number = _next_version_number(database, channel, date)
        document = {
            "version_id": str(uuid.uuid4()),
            "channel": channel,
            "date": date,
            "file_id": file_id,
            "job_id": str(job_id or ""),
            "genre": genre or "",
            "version_number": version_number,
            "created_at": utc_now_iso(),
            "reason": reason,
            "modification_count": len(modifications or []),
            "modifications_snapshot": modifications or [],
            "original_filename": original_filename,
            "xlsx_data": Binary(bytes(xlsx_bytes)),
            "source_version": source_version,
            **metadata,
        }
        try:
            database[VERSION_COLLECTION].insert_one(document)
            return document
        except DuplicateKeyError:
            if attempt == 4:
                raise
    raise RuntimeError("Unable to allocate a report version number")


def version_summary(document):
    return {
        key: document.get(key)
        for key in (
            "version_id",
            "channel",
            "date",
            "file_id",
            "job_id",
            "version_number",
            "created_at",
            "reason",
            "modification_count",
            "source_version",
            "original_filename",
            "workbook_size_bytes",
            "sheet_count",
            "rows_processed",
            "checksum_sha256",
        )
    }


def _duration_seconds(value):
    if value is None:
        return 0
    if hasattr(value, "total_seconds"):
        return int(value.total_seconds())
    text = str(value).strip()
    parts = text.split(":")
    if len(parts) != 3:
        return 0
    try:
        hours, minutes, seconds = (int(float(part)) for part in parts)
        return hours * 3600 + minutes * 60 + seconds
    except (TypeError, ValueError):
        return 0


def _number(value):
    try:
        return int(float(str(value or "0").replace(",", "").strip()))
    except (TypeError, ValueError):
        return 0


def _summary_values(rows):
    if len(rows) < 3:
        return {}
    headers = [str(value or "").strip().upper() for value in rows[1]]
    values = rows[2]
    return {
        header: values[index] if index < len(values) else ""
        for index, header in enumerate(headers)
        if header
    }


def _lookup_metric(values, keyword):
    keyword = keyword.upper()
    for header, value in values.items():
        if keyword in header:
            return value
    return 0


def validate_report_consistency(database, channel, date, job_id=""):
    timestamp = utc_now_iso()
    checks = []
    failures = []

    def check(name, expected, actual):
        passed = expected == actual
        result = {
            "name": name,
            "passed": passed,
            "expected": expected,
            "actual": actual,
        }
        checks.append(result)
        if not passed:
            failures.append(result)

    processed = database.processed_files.find_one(
        {"channel_name": channel, "date": date}
    )
    if not processed:
        check("stored_report_exists", True, False)
    else:
        xlsx_bytes = bytes(processed.get("xlsx_data") or b"")
        try:
            metadata = workbook_metadata(xlsx_bytes)
            stored_sheets = list(
                database.sheets.find(
                    {"channel_name": channel, "date": date},
                    {"_id": 0, "sheet_name": 1, "row_count": 1, "rows": 1, "file_id": 1},
                )
            )
            check("sheet_count", metadata["sheet_count"], len(stored_sheets))
            for sheet in stored_sheets:
                sheet_name = sheet.get("sheet_name", "")
                check(
                    f"row_count:{sheet_name}",
                    metadata["sheet_rows"].get(sheet_name, 0),
                    int(sheet.get("row_count") or 0),
                )
                check(
                    f"file_id:{sheet_name}",
                    processed.get("file_id"),
                    sheet.get("file_id"),
                )

            workbook = load_workbook(
                io.BytesIO(xlsx_bytes),
                read_only=True,
                data_only=True,
            )
            try:
                workbook_rows = {
                    worksheet.title: [
                        [str(value) if value is not None else "" for value in row]
                        for row in worksheet.iter_rows(values_only=True)
                    ]
                    for worksheet in workbook.worksheets
                }
            finally:
                workbook.close()

            for sheet in stored_sheets:
                sheet_name = sheet.get("sheet_name", "")
                check(
                    f"workbook_matches_stored_sheet:{sheet_name}",
                    workbook_rows.get(sheet_name, []),
                    sheet.get("rows") or [],
                )

            summary = next(
                (
                    sheet
                    for sheet in stored_sheets
                    if sheet.get("sheet_name") == "TABSONS SUMMARY"
                ),
                None,
            )
            if summary:
                values = _summary_values(summary.get("rows") or [])
                workbook_summary_values = _summary_values(
                    workbook_rows.get("TABSONS SUMMARY", [])
                )
                check(
                    "dashboard_totals_match_workbook",
                    workbook_summary_values,
                    values,
                )
                for source in ("TABSONS", "BARC"):
                    total_count = _number(_lookup_metric(values, f"{source} LINE ITEM"))
                    component_count = sum(
                        _number(_lookup_metric(values, f"{source} {label} COUNT"))
                        for label in ("COMMERCIAL", "PROMO", "PROMO SPONSOR", "PROGRAM")
                    )
                    component_count += _number(
                        _lookup_metric(
                            values,
                            "TABSONS ICA (COUNT)"
                            if source == "TABSONS"
                            else "BARC ICA COUNT",
                        )
                    )
                    check(f"{source.lower()}_count_total", total_count, component_count)

                    total_duration = _duration_seconds(
                        _lookup_metric(values, f"{source} DURATION")
                    )
                    component_duration = sum(
                        (
                            _duration_seconds(
                                _lookup_metric(
                                    values,
                                    f"{source} PROMO SPONSOR COUNT DURATION",
                                )
                            )
                            if label == "PROMO SPONSOR"
                            else _duration_seconds(
                                _lookup_metric(values, f"{source} {label} DURATION")
                            )
                        )
                        for label in ("COMMERCIAL", "PROMO", "PROMO SPONSOR", "PROGRAM")
                    )
                    component_duration += _duration_seconds(
                        _lookup_metric(
                            values,
                            "TABSONS ICA (DURATION)"
                            if source == "TABSONS"
                            else "BARC ICA DURATION",
                        )
                    )
                    check(
                        f"{source.lower()}_duration_total",
                        total_duration,
                        component_duration,
                    )

            latest_version = database[VERSION_COLLECTION].find_one(
                {"channel": channel, "date": date},
                sort=[("version_number", DESCENDING)],
            )
            if latest_version:
                check(
                    "version_source_checksum",
                    metadata["checksum_sha256"],
                    latest_version.get("checksum_sha256"),
                )
                current_modifications = list(
                    database.brand_modifications.find(
                        {"channel_name": channel, "date": date},
                        {"_id": 0},
                    ).sort("timestamp", ASCENDING)
                )
                check(
                    "commercial_modification_count",
                    len(current_modifications),
                    int(latest_version.get("modification_count") or 0),
                )
        except Exception as exc:
            failures.append(
                {
                    "name": "validation_execution",
                    "passed": False,
                    "expected": "valid workbook and stored data",
                    "actual": f"{type(exc).__name__}: {exc}",
                }
            )
            checks.extend(failures[-1:])

    status = "PASSED" if not failures else "FAILED"
    document = {
        "validation_id": str(uuid.uuid4()),
        "timestamp": timestamp,
        "channel": channel,
        "date": date,
        "job_id": str(job_id or ""),
        "status": status,
        "checks": checks,
        "failure_count": len(failures),
        "failures": failures,
    }
    database[VALIDATION_COLLECTION].insert_one(document)
    if failures:
        record_audit(
            database,
            "VALIDATION_FAILURE",
            "FAILED",
            channel=channel,
            date=date,
            job_id=job_id,
            details={"validation_id": document["validation_id"], "failures": failures},
        )
    return document


def restore_report_version(
    database,
    channel,
    date,
    version_number,
    parse_workbook_sheets,
    job_id="",
):
    version = database[VERSION_COLLECTION].find_one(
        {
            "channel": channel,
            "date": date,
            "version_number": int(version_number),
        }
    )
    if not version:
        raise LookupError("Report version not found")

    xlsx_bytes = bytes(version.get("xlsx_data") or b"")
    if workbook_checksum(xlsx_bytes) != version.get("checksum_sha256"):
        raise ValueError("Stored report version checksum validation failed")

    new_file_id = str(uuid.uuid4())
    restored_at = utc_now_iso()
    genre = version.get("genre") or ""
    sheet_documents = parse_workbook_sheets(
        xlsx_bytes,
        new_file_id,
        channel,
        date,
        genre,
    )
    modifications = [
        {key: value for key, value in document.items() if key != "_id"}
        for document in (version.get("modifications_snapshot") or [])
    ]
    lookup = {"channel_name": channel, "date": date}
    backup_report = database.processed_files.find_one(lookup)
    backup_sheets = list(database.sheets.find(lookup))
    backup_modifications = list(database.brand_modifications.find(lookup))
    restored_version = None

    try:
        database.sheets.delete_many(lookup)
        database.processed_files.replace_one(
            lookup,
            {
                "file_id": new_file_id,
                "channel_name": channel,
                "genre": genre,
                "date": date,
                "original_filename": version.get("original_filename") or "report.xlsx",
                "xlsx_data": Binary(xlsx_bytes),
                "uploaded_at": restored_at,
                "restored_from_version": int(version_number),
                "workbook_size_bytes": len(xlsx_bytes),
                "sheet_count": len(sheet_documents),
                "rows_processed": sum(
                    int(document.get("row_count") or 0)
                    for document in sheet_documents
                ),
            },
            upsert=True,
        )
        if sheet_documents:
            database.sheets.insert_many(sheet_documents)
        database.brand_modifications.delete_many(lookup)
        if modifications:
            database.brand_modifications.insert_many(
                [dict(document) for document in modifications]
            )

        restored_version = create_report_version(
            database,
            channel,
            date,
            xlsx_bytes,
            version.get("original_filename") or "report.xlsx",
            new_file_id,
            reason=f"RESTORE_VERSION_{version_number}",
            modifications=modifications,
            job_id=job_id,
            genre=genre,
            source_version=int(version_number),
        )
        database.processed_files.update_one(
            lookup,
            {
                "$set": {
                    "active_version_number": restored_version["version_number"],
                }
            },
        )
        return restored_version
    except Exception:
        database.sheets.delete_many(lookup)
        database.brand_modifications.delete_many(lookup)
        if backup_report:
            database.processed_files.replace_one(lookup, backup_report, upsert=True)
        else:
            database.processed_files.delete_one(lookup)
        if backup_sheets:
            database.sheets.insert_many(backup_sheets)
        if backup_modifications:
            database.brand_modifications.insert_many(backup_modifications)
        if restored_version:
            database[VERSION_COLLECTION].delete_one(
                {"version_id": restored_version["version_id"]}
            )
        raise
