import io
import unittest
import uuid
from copy import deepcopy
from types import SimpleNamespace
from unittest.mock import patch

from bson.binary import Binary
from openpyxl import Workbook, load_workbook

import app
from audit_recovery import (
    AUDIT_COLLECTION,
    VERSION_COLLECTION,
    create_report_version,
    record_audit,
    restore_report_version,
    validate_report_consistency,
)


def matches(document, query):
    for key, expected in (query or {}).items():
        actual = document.get(key)
        if isinstance(expected, dict) and "$in" in expected:
            if actual not in expected["$in"]:
                return False
        elif actual != expected:
            return False
    return True


def project(document, projection):
    if not projection:
        return deepcopy(document)
    included = [key for key, value in projection.items() if value and key != "_id"]
    if included:
        result = {key: deepcopy(document.get(key)) for key in included if key in document}
    else:
        result = deepcopy(document)
    for key, value in projection.items():
        if not value:
            result.pop(key, None)
    return result


class FakeCursor:
    def __init__(self, documents):
        self.documents = documents

    def sort(self, key, direction=None):
        sort_fields = key if isinstance(key, list) else [(key, direction)]
        for field, order in reversed(sort_fields):
            self.documents.sort(
                key=lambda document: document.get(field, ""),
                reverse=order == -1,
            )
        return self

    def limit(self, count):
        self.documents = self.documents[:count]
        return self

    def __iter__(self):
        return iter(self.documents)


class FakeCollection:
    def __init__(self):
        self.documents = []

    def create_index(self, *args, **kwargs):
        return kwargs.get("name", "index")

    def insert_one(self, document):
        stored = deepcopy(document)
        stored.setdefault("_id", str(uuid.uuid4()))
        document.setdefault("_id", stored["_id"])
        self.documents.append(stored)
        return SimpleNamespace(inserted_id=stored["_id"])

    def insert_many(self, documents):
        ids = [self.insert_one(document).inserted_id for document in documents]
        return SimpleNamespace(inserted_ids=ids)

    def find_one(self, query, projection=None, sort=None):
        documents = [document for document in self.documents if matches(document, query)]
        if sort:
            documents = list(FakeCursor(documents).sort(sort).documents)
        return project(documents[0], projection) if documents else None

    def find(self, query=None, projection=None):
        return FakeCursor(
            [
                project(document, projection)
                for document in self.documents
                if matches(document, query or {})
            ]
        )

    def count_documents(self, query):
        return sum(matches(document, query) for document in self.documents)

    def delete_many(self, query):
        before = len(self.documents)
        self.documents = [
            document for document in self.documents if not matches(document, query)
        ]
        return SimpleNamespace(deleted_count=before - len(self.documents))

    def delete_one(self, query):
        for index, document in enumerate(self.documents):
            if matches(document, query):
                self.documents.pop(index)
                return SimpleNamespace(deleted_count=1)
        return SimpleNamespace(deleted_count=0)

    def replace_one(self, query, replacement, upsert=False):
        for index, document in enumerate(self.documents):
            if matches(document, query):
                stored = deepcopy(replacement)
                stored.setdefault("_id", document.get("_id", str(uuid.uuid4())))
                self.documents[index] = stored
                return SimpleNamespace(modified_count=1)
        if upsert:
            self.insert_one(replacement)
            return SimpleNamespace(modified_count=0, upserted_id=replacement.get("_id"))
        return SimpleNamespace(modified_count=0)

    def update_one(self, query, update, upsert=False):
        document = self.find_one(query)
        if not document and upsert:
            document = deepcopy(query)
            document.update(update.get("$setOnInsert", {}))
            document.update(update.get("$set", {}))
            self.insert_one(document)
            return SimpleNamespace(modified_count=0)
        if not document:
            return SimpleNamespace(modified_count=0)
        for stored in self.documents:
            if stored.get("_id") == document.get("_id"):
                stored.update(deepcopy(update.get("$set", {})))
                return SimpleNamespace(modified_count=1)
        return SimpleNamespace(modified_count=0)


class FakeDatabase:
    def __init__(self):
        self.collections = {}

    def __getitem__(self, name):
        return self.collections.setdefault(name, FakeCollection())

    def __getattr__(self, name):
        if name.startswith("_"):
            raise AttributeError(name)
        return self[name]


def workbook_bytes(total=3):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "TABSONS SUMMARY"
    worksheet.append(["SUMMARY"])
    worksheet.append(
        [
            "TABSONS LINE ITEM",
            "TABSONS DURATION",
            "TABSONS COMMERCIAL COUNT",
            "TABSONS COMMERCIAL DURATION",
            "TABSONS PROMO COUNT",
            "TABSONS PROMO DURATION",
            "TABSONS PROMO SPONSOR COUNT",
            "TABSONS PROMO SPONSOR COUNT DURATION",
            "TABSONS PROGRAM COUNT",
            "TABSONS PROGRAM DURATION",
            "BARC LINE ITEM",
            "BARC DURATION",
            "BARC COMMERCIAL COUNT",
            "BARC COMMERCIAL DURATION",
            "BARC PROMO COUNT",
            "BARC PROMO DURATION",
            "BARC PROMO SPONSOR COUNT",
            "BARC PROMO SPONSOR COUNT DURATION",
            "BARC PROGRAM COUNT",
            "BARC PROGRAM DURATION",
        ]
    )
    worksheet.append(
        [
            total,
            "00:00:30",
            1,
            "00:00:10",
            1,
            "00:00:10",
            0,
            "00:00:00",
            1,
            "00:00:10",
            total,
            "00:00:30",
            1,
            "00:00:10",
            1,
            "00:00:10",
            0,
            "00:00:00",
            1,
            "00:00:10",
        ]
    )
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def parse_sheets(xlsx_bytes, file_id, channel, date, genre, timing_stats=None):
    workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    try:
        documents = []
        for worksheet in workbook.worksheets:
            rows = [
                [str(value) if value is not None else "" for value in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
            documents.append(
                {
                    "file_id": file_id,
                    "channel_name": channel,
                    "genre": genre,
                    "date": date,
                    "sheet_name": worksheet.title,
                    "headers": rows[0],
                    "rows": rows,
                    "row_count": len(rows),
                    "col_count": len(rows[0]),
                }
            )
        return documents
    finally:
        workbook.close()


class AuditRecoveryTests(unittest.TestCase):
    def setUp(self):
        self.database = FakeDatabase()
        self.xlsx = workbook_bytes()

    def store_report(self, file_id="file-1"):
        self.database.processed_files.insert_one(
            {
                "file_id": file_id,
                "channel_name": "CHANNEL",
                "genre": "NEWS",
                "date": "10/05/2026",
                "original_filename": "report.xlsx",
                "xlsx_data": Binary(self.xlsx),
            }
        )
        self.database.sheets.insert_many(
            parse_sheets(
                self.xlsx,
                file_id,
                "CHANNEL",
                "10/05/2026",
                "NEWS",
            )
        )

    def test_audit_event_contains_required_fields(self):
        event = record_audit(
            self.database,
            "FILE_UPLOAD",
            "QUEUED",
            channel="CHANNEL",
            date="10/05/2026",
            job_id="job-1",
            details={"filename": "input.xlsx"},
        )

        for field in (
            "event_id",
            "timestamp",
            "user_action",
            "channel",
            "date",
            "job_id",
            "status",
            "details",
        ):
            self.assertIn(field, event)
        self.assertEqual(self.database[AUDIT_COLLECTION].count_documents({}), 1)

    def test_versions_are_append_only_and_numbered(self):
        first = create_report_version(
            self.database,
            "CHANNEL",
            "10/05/2026",
            self.xlsx,
            "report.xlsx",
            "file-1",
            "ORIGINAL_REPORT",
        )
        second = create_report_version(
            self.database,
            "CHANNEL",
            "10/05/2026",
            self.xlsx,
            "report.xlsx",
            "file-1",
            "COMMERCIAL_MERGE",
            modifications=[{"action": "merge_to_matched"}],
        )

        self.assertEqual(first["version_number"], 1)
        self.assertEqual(second["version_number"], 2)
        self.assertEqual(second["modification_count"], 1)
        self.assertEqual(self.database[VERSION_COLLECTION].count_documents({}), 2)

    def test_validation_detects_stored_sheet_mismatch(self):
        self.store_report()
        create_report_version(
            self.database,
            "CHANNEL",
            "10/05/2026",
            self.xlsx,
            "report.xlsx",
            "file-1",
            "ORIGINAL_REPORT",
        )
        passed = validate_report_consistency(
            self.database,
            "CHANNEL",
            "10/05/2026",
        )
        self.assertEqual(passed["status"], "PASSED")

        self.database.sheets.documents[0]["row_count"] = 999
        failed = validate_report_consistency(
            self.database,
            "CHANNEL",
            "10/05/2026",
        )
        self.assertEqual(failed["status"], "FAILED")
        self.assertTrue(
            any(failure["name"].startswith("row_count:") for failure in failed["failures"])
        )

    def test_specific_version_restore_preserves_history(self):
        self.store_report(file_id="current-file")
        first = create_report_version(
            self.database,
            "CHANNEL",
            "10/05/2026",
            self.xlsx,
            "report.xlsx",
            "original-file",
            "ORIGINAL_REPORT",
        )
        self.database.brand_modifications.insert_one(
            {
                "channel_name": "CHANNEL",
                "date": "10/05/2026",
                "action": "merge_to_matched",
                "timestamp": "2026-06-12T10:00:00",
            }
        )
        create_report_version(
            self.database,
            "CHANNEL",
            "10/05/2026",
            self.xlsx,
            "report.xlsx",
            "current-file",
            "COMMERCIAL_MERGE",
            modifications=[
                {
                    "channel_name": "CHANNEL",
                    "date": "10/05/2026",
                    "action": "merge_to_matched",
                    "timestamp": "2026-06-12T10:00:00",
                }
            ],
        )

        restored = restore_report_version(
            self.database,
            "CHANNEL",
            "10/05/2026",
            first["version_number"],
            parse_sheets,
        )

        self.assertEqual(restored["version_number"], 3)
        self.assertEqual(restored["source_version"], 1)
        self.assertEqual(
            self.database.brand_modifications.count_documents(
                {"channel_name": "CHANNEL", "date": "10/05/2026"}
            ),
            0,
        )
        current = self.database.processed_files.find_one(
            {"channel_name": "CHANNEL", "date": "10/05/2026"}
        )
        self.assertEqual(current["active_version_number"], 3)
        self.assertEqual(bytes(current["xlsx_data"]), self.xlsx)

    def test_admin_diagnostics_require_key(self):
        record_audit(self.database, "FILE_UPLOAD", "QUEUED")
        with (
            patch.object(app, "ADMIN_API_KEY", "secret"),
            patch.object(app, "get_mongo_database", return_value=self.database),
        ):
            with app.app.test_client() as client:
                denied = client.get("/api/admin/audit-logs")
                allowed = client.get(
                    "/api/admin/audit-logs",
                    headers={"X-Admin-Key": "secret"},
                )

        self.assertEqual(denied.status_code, 401)
        self.assertEqual(allowed.status_code, 200)
        self.assertEqual(len(allowed.get_json()), 1)


if __name__ == "__main__":
    unittest.main()
