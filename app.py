"""
Content Analyzer Flask backend with MongoDB Atlas storage.

Serves the dashboard frontend and provides APIs for file processing,
sheet retrieval, brand management, and report downloads.
"""

import io
import json
import logging
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import uuid
import zipfile
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path
import pandas as pd
from bson import BSON
from bson.decimal128 import Decimal128
from bson.binary import Binary
from bson.objectid import ObjectId
from dotenv import load_dotenv
from flask import Flask, jsonify, request, send_file, send_from_directory
from flask_cors import CORS
from gridfs import GridFSBucket, NoFile
from openpyxl import load_workbook
from werkzeug.exceptions import HTTPException
from pymongo import ASCENDING, MongoClient, ReturnDocument
from pymongo.errors import InvalidURI, PyMongoError

load_dotenv()

MONGO_URI = os.getenv("MONGO_URI")
DB_NAME = os.getenv("DB_NAME") or os.getenv("MONGO_DB_NAME", "content_analyzer")
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
SUPPORTED_GENRES = ("NEWS", "GEC / ENTERTAINMENT", "SPORTS", "MUSIC")
CHANNEL_STATUSES = ("active", "pending_assignment", "inactive")
CONFIGURED_DEFAULT_CHANNEL_GENRE = os.getenv("DEFAULT_CHANNEL_GENRE", "NEWS").strip().upper()
DEFAULT_CHANNEL_GENRE = (
    CONFIGURED_DEFAULT_CHANNEL_GENRE
    if CONFIGURED_DEFAULT_CHANNEL_GENRE in SUPPORTED_GENRES
    else "NEWS"
)

BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_FILE = BASE_DIR / "brand_comparison_template.xlsx"
COMPARISON_SCRIPT = BASE_DIR / "barc_nct_comparison.py"
JOB_STATUSES = ("QUEUED", "PROCESSING", "COMPLETED", "FAILED")
JOB_UPLOAD_RETENTION_DAYS = int(os.getenv("JOB_UPLOAD_RETENTION_DAYS", "7"))
JOB_LEASE_SECONDS = int(os.getenv("JOB_LEASE_SECONDS", "600"))
MAX_UPLOAD_BYTES = int(os.getenv("MAX_UPLOAD_BYTES", str(100 * 1024 * 1024)))
COMPARISON_PROCESS_TIMEOUT_SECONDS = int(
    os.getenv("COMPARISON_PROCESS_TIMEOUT_SECONDS", "1800")
)

app = Flask(__name__, static_folder=".", static_url_path="")
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-secret-key-change-me")
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_BYTES
CORS(app)

logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
)
app.logger.setLevel(getattr(logging, LOG_LEVEL, logging.INFO))

_mongo_client = None
_mongo_db = None
_indexes_ready = False
_genre_alignment_ready = False


def peak_rss_mib():
    """Return peak resident memory where the platform exposes it."""
    try:
        import resource

        peak = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss
        if sys.platform == "darwin":
            peak /= 1024
        return round(peak / 1024, 2)
    except (ImportError, AttributeError, OSError):
        return None


def log_stage(request_id: str, stage: str, started: float, **details) -> float:
    elapsed_ms = (time.perf_counter() - started) * 1000
    fields = {
        "request_id": request_id,
        "stage": stage,
        "duration_ms": round(elapsed_ms, 1),
        **details,
    }
    peak = peak_rss_mib()
    if peak is not None:
        fields["peak_rss_mib"] = peak
    app.logger.info(
        "Processing timing %s",
        " ".join(f"{key}={value}" for key, value in fields.items()),
    )
    return elapsed_ms


def is_api_request() -> bool:
    return request.path.startswith("/api/") or request.path in {"/analyze", "/healthz"}


def to_json_safe(value):
    """Recursively convert MongoDB/BSON and pandas values into JSON-safe data."""
    if isinstance(value, ObjectId):
        return str(value)
    if isinstance(value, Decimal128):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, Binary):
        return f"<binary {len(value)} bytes>"
    if isinstance(value, (bytes, bytearray, memoryview)):
        return f"<binary {len(value)} bytes>"
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {str(to_json_safe(key)): to_json_safe(val) for key, val in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [to_json_safe(item) for item in value]

    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    if hasattr(value, "item"):
        try:
            return value.item()
        except (TypeError, ValueError):
            pass

    return value


def validate_json_payload(payload, route_name: str = "") -> None:
    """Fail fast if an API payload still contains non-serializable objects."""
    try:
        json.dumps(payload)
    except TypeError as exc:
        app.logger.exception("JSON validation failed for %s: %s", route_name or request.path, exc)
        raise


def json_response(payload, status: int = 200):
    safe_payload = to_json_safe(payload)
    validate_json_payload(safe_payload, request.endpoint or request.path)
    response = jsonify(safe_payload)
    response.status_code = status
    return response


def api_error(message: str, status: int = 500, **extra):
    payload = {"success": False, "error": str(message)}
    payload.update(extra)
    return json_response(payload, status)


def utc_now_iso() -> str:
    return datetime.utcnow().isoformat()


def iso_after(seconds: int = 0, days: int = 0) -> str:
    return (datetime.utcnow() + timedelta(seconds=seconds, days=days)).isoformat()


def normalize_channel_name(channel_name: str) -> str:
    return re.sub(r"\s+", " ", str(channel_name or "")).strip()


def channel_key(channel_name: str) -> str:
    return normalize_channel_name(channel_name).upper()


def validate_genre(genre: str) -> str:
    normalized = str(genre or "").strip().upper()
    if normalized not in SUPPORTED_GENRES:
        raise ValueError(f"genre must be one of: {', '.join(SUPPORTED_GENRES)}")
    return normalized


def validate_channel_status(status: str) -> str:
    normalized = str(status or "").strip().lower()
    if normalized not in CHANNEL_STATUSES:
        raise ValueError(f"status must be one of: {', '.join(CHANNEL_STATUSES)}")
    return normalized


def get_mongo_client() -> MongoClient:
    """Return the shared MongoDB client with connection pooling enabled."""
    global _mongo_client
    if _mongo_client is None:
        if not MONGO_URI:
            raise RuntimeError("MONGO_URI environment variable is required")

        try:
            _mongo_client = MongoClient(
                MONGO_URI,
                appname="content-analyzer",
                maxPoolSize=int(os.getenv("MONGO_MAX_POOL_SIZE", "50")),
                minPoolSize=int(os.getenv("MONGO_MIN_POOL_SIZE", "0")),
                serverSelectionTimeoutMS=int(os.getenv("MONGO_SERVER_SELECTION_TIMEOUT_MS", "120000  ")),
                connectTimeoutMS=int(os.getenv("MONGO_CONNECT_TIMEOUT_MS", "120000")),
                socketTimeoutMS=int(os.getenv("MONGO_SOCKET_TIMEOUT_MS", "120000 ")),
                retryWrites=True,
            )
        except InvalidURI as exc:
            raise RuntimeError(
                "Invalid MONGO_URI. URL-encode the MongoDB username/password "
                "in the Atlas URI, for example @ as %40, # as %23, / as %2F."
            ) from exc
    return _mongo_client


def get_mongo_database():
    """Return the configured MongoDB database."""
    global _mongo_db
    if _mongo_db is None:
        _mongo_db = get_mongo_client()[DB_NAME]
    return _mongo_db


def validate_mongo_connection() -> None:
    """Validate MongoDB Atlas connectivity with a ping command."""
    started = time.perf_counter()
    app.logger.debug("Validating MongoDB connection for database '%s'", DB_NAME)
    get_mongo_client().admin.command("ping")
    app.logger.info(
        "MongoDB connection validated database=%s duration_ms=%.1f",
        DB_NAME,
        (time.perf_counter() - started) * 1000,
    )


def ensure_mongo_indexes() -> None:
    """Create indexes required by the dashboard lookup patterns."""
    global _indexes_ready
    if _indexes_ready:
        return

    started = time.perf_counter()
    validate_mongo_connection()
    database = get_mongo_database()
    database.processed_files.create_index(
        [("channel_name", ASCENDING), ("date", ASCENDING)],
        name="idx_pf_lookup",
        unique=True,
    )
    database.processed_files.create_index(
        [("file_id", ASCENDING)],
        name="idx_pf_file_id",
        unique=True,
    )
    database.sheets.create_index(
        [("channel_name", ASCENDING), ("date", ASCENDING), ("sheet_name", ASCENDING)],
        name="idx_sheets_lookup",
        unique=True,
    )
    database.sheets.create_index(
        [("file_id", ASCENDING)],
        name="idx_sheets_file_id",
    )
    database.brand_modifications.create_index(
        [("channel_name", ASCENDING), ("date", ASCENDING), ("timestamp", ASCENDING)],
        name="idx_bm_lookup",
    )
    database.genre_channel_master.create_index(
        [("channel_key", ASCENDING)],
        name="idx_gcm_channel_key",
        unique=True,
        partialFilterExpression={"channel_key": {"$type": "string"}},
    )
    database.genre_channel_master.create_index(
        [("genre", ASCENDING), ("status", ASCENDING), ("channel_name", ASCENDING)],
        name="idx_gcm_genre_channels",
    )
    database.processing_jobs.create_index(
        [("job_id", ASCENDING)],
        name="idx_jobs_job_id",
        unique=True,
    )
    database.processing_jobs.create_index(
        [("status", ASCENDING), ("created_at", ASCENDING)],
        name="idx_jobs_queue",
    )
    database.processing_jobs.create_index(
        [("status", ASCENDING), ("lease_expires_at", ASCENDING)],
        name="idx_jobs_recovery",
    )
    database.processing_jobs.create_index(
        [("input_expires_at", ASCENDING)],
        name="idx_jobs_cleanup",
    )
    _indexes_ready = True
    app.logger.info(
        "MongoDB indexes are ready duration_ms=%.1f",
        (time.perf_counter() - started) * 1000,
    )


def get_collections():
    ensure_mongo_indexes()
    database = get_mongo_database()
    return (
        database.processed_files,
        database.sheets,
        database.brand_modifications,
    )


def get_channel_master_collection():
    ensure_mongo_indexes()
    return get_mongo_database().genre_channel_master


def get_processing_jobs_collection():
    ensure_mongo_indexes()
    return get_mongo_database().processing_jobs


def get_job_files_bucket():
    ensure_mongo_indexes()
    return GridFSBucket(get_mongo_database(), bucket_name="job_files")


def ensure_channel_record(
    channel_name: str,
    genre: str = None,
    status: str = None,
):
    """Return a valid master record, creating a pending assignment when needed."""
    normalized_name = normalize_channel_name(channel_name)
    normalized_key = channel_key(normalized_name)
    if not normalized_key:
        raise ValueError("channel_name is required")

    master = get_channel_master_collection()
    existing = master.find_one({"channel_key": normalized_key})
    if not existing:
        existing = master.find_one({"channel_name": normalized_name})
    now = utc_now_iso()

    if existing:
        updates = {}
        existing_genre = existing.get("genre")
        existing_status = existing.get("status")

        if genre is not None:
            updates["genre"] = validate_genre(genre)
        elif existing_genre not in SUPPORTED_GENRES:
            updates["genre"] = DEFAULT_CHANNEL_GENRE

        if status is not None:
            updates["status"] = validate_channel_status(status)
        elif existing_status not in CHANNEL_STATUSES:
            updates["status"] = "pending_assignment"

        if not existing.get("channel_name"):
            updates["channel_name"] = normalized_name
        if not existing.get("created_at"):
            updates["created_at"] = now
        if not existing.get("updated_at"):
            updates["updated_at"] = now
        if not existing.get("channel_key"):
            updates["channel_key"] = normalized_key

        if updates:
            if set(updates) != {"updated_at"}:
                updates["updated_at"] = now
            master.update_one({"_id": existing["_id"]}, {"$set": updates})
            existing.update(updates)
        return existing

    assigned_genre = validate_genre(genre or DEFAULT_CHANNEL_GENRE)
    assigned_status = validate_channel_status(status or "pending_assignment")
    document = {
        "channel_key": normalized_key,
        "channel_name": normalized_name,
        "genre": assigned_genre,
        "status": assigned_status,
        "created_at": now,
        "updated_at": now,
    }
    master.update_one(
        {"channel_key": normalized_key},
        {"$setOnInsert": document},
        upsert=True,
    )
    return master.find_one({"channel_key": normalized_key})


def migrate_genre_hierarchy(force: bool = False):
    """Align legacy channel documents with the genre/channel master collection."""
    global _genre_alignment_ready
    if _genre_alignment_ready and not force:
        return {"channels": 0, "processed_files": 0, "sheets": 0, "already_aligned": True}

    processed_files, sheets, _ = get_collections()
    channel_names = [
        name for name in processed_files.distinct("channel_name")
        if normalize_channel_name(name)
    ]
    migrated_files = 0
    migrated_sheets = 0

    for name in channel_names:
        record = ensure_channel_record(name)
        genre = record["genre"]
        migrated_files += processed_files.update_many(
            {"channel_name": name, "genre": {"$ne": genre}},
            {"$set": {"genre": genre}},
        ).modified_count
        migrated_sheets += sheets.update_many(
            {"channel_name": name, "genre": {"$ne": genre}},
            {"$set": {"genre": genre}},
        ).modified_count

    _genre_alignment_ready = True
    result = {
        "channels": len(channel_names),
        "processed_files": migrated_files,
        "sheets": migrated_sheets,
        "already_aligned": False,
    }
    app.logger.info("Genre hierarchy migration complete: %s", result)
    return result


def update_channel_assignment(channel_name: str, genre: str, status: str = "active"):
    """Assign a validated genre and synchronize existing report documents."""
    record = ensure_channel_record(channel_name, genre=genre, status=status)
    processed_files, sheets, _ = get_collections()
    canonical_name = record["channel_name"]
    processed_files.update_many(
        {"channel_name": canonical_name},
        {"$set": {"genre": record["genre"]}},
    )
    sheets.update_many(
        {"channel_name": canonical_name},
        {"$set": {"genre": record["genre"]}},
    )
    return record


def database_error_response(exc):
    app.logger.exception("Database operation failed: %s", exc)
    return api_error("Database operation failed", 500, details=str(exc))


@app.before_request
def log_request_start():
    request._started_at = time.time()
    if is_api_request():
        file_names = [storage.filename for storage in request.files.values()]
        json_keys = []
        if request.is_json:
            body = request.get_json(silent=True) or {}
            json_keys = list(body.keys()) if isinstance(body, dict) else []
        app.logger.info(
            "API request start method=%s path=%s args=%s json_keys=%s files=%s",
            request.method,
            request.path,
            dict(request.args),
            json_keys,
            file_names,
        )


@app.after_request
def log_request_finish(response):
    if is_api_request():
        elapsed_ms = int((time.time() - getattr(request, "_started_at", time.time())) * 1000)
        app.logger.info(
            "API request finish method=%s path=%s status=%s mimetype=%s elapsed_ms=%s",
            request.method,
            request.path,
            response.status_code,
            response.mimetype,
            elapsed_ms,
        )
    return response


@app.errorhandler(Exception)
def handle_exception(e):
    print("GLOBAL ERROR:", str(e))
    status = e.code if isinstance(e, HTTPException) else 500
    if status >= 500:
        app.logger.exception("Unhandled exception on %s %s: %s", request.method, request.path, e)
    else:
        app.logger.warning("HTTP exception on %s %s: %s", request.method, request.path, e)
    return json_response({"success": False, "error": str(e)}, status)


def route_guard(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        try:
            app.logger.debug("Entering route endpoint=%s path=%s", request.endpoint, request.path)
            return func(*args, **kwargs)
        except Exception as exc:
            return handle_exception(exc)

    wrapper._route_guarded = True
    return wrapper


def fmt_size(b: int) -> str:
    if b < 1024:
        return f"{b} B"
    if b < 1024**2:
        return f"{b / 1024:.1f} KB"
    return f"{b / 1024**2:.1f} MB"


def build_filename(channel: str, date_str: str) -> str:
    channel_clean = re.sub(r"[^A-Z0-9]", "", str(channel).upper().strip())
    if not channel_clean:
        channel_clean = "UNKNOWN"

    match = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", str(date_str).strip())
    if match:
        date_clean = f"{match.group(1)}{match.group(2)}{match.group(3)}"
    else:
        iso_match = re.match(r"^(\d{4})-(\d{2})-(\d{2})", str(date_str).strip())
        date_clean = (
            f"{iso_match.group(3)}{iso_match.group(2)}{iso_match.group(1)}"
            if iso_match
            else "00000000"
        )

    return f"{channel_clean}({date_clean}) barc_nct_comparison"


def run_comparison(file_bytes: bytes, original_name: str, progress_callback=None) -> tuple:
    """Run barc_nct_comparison.py and return (xlsx_bytes, output_filename, stats)."""
    total_started = time.perf_counter()
    output_filename = "output.xlsx"
    stats = {"timings_ms": {}, "input_bytes": len(file_bytes)}

    if progress_callback:
        progress_callback(10, "Reading Workbook")

    metadata_started = time.perf_counter()
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
        source_series = df.get("source", pd.Series([""] * len(df)))
        normalized_source = source_series.fillna("").astype(str).str.upper().str.strip()
        df_barc = df[normalized_source == "BARC XML"]

        if len(df_barc):
            channel = str(df_barc["channel name"].iloc[0])
            date_val = str(df_barc["TelecastDate"].iloc[0])
            output_filename = build_filename(channel, date_val) + ".xlsx"
            stats["channel"] = channel
            stats["date"] = date_val
            stats["barc_rows"] = int(len(df_barc))
            stats["nct_rows"] = int((normalized_source == "NCT").sum())
    except Exception as exc:
        stats["metadata_error"] = str(exc)
    finally:
        stats["timings_ms"]["metadata_read"] = round(
            (time.perf_counter() - metadata_started) * 1000,
            1,
        )

    with tempfile.TemporaryDirectory() as tmpdir:
        setup_started = time.perf_counter()
        input_path = Path(tmpdir) / "brand_comparison_template.xlsx"
        input_path.write_bytes(file_bytes)

        temp_script = Path(tmpdir) / "barc_nct_comparison.py"
        shutil.copy(COMPARISON_SCRIPT, temp_script)
        progress_path = Path(tmpdir) / "comparison_progress.json"
        stats["timings_ms"]["temporary_setup"] = round(
            (time.perf_counter() - setup_started) * 1000,
            1,
        )
        child_env = os.environ.copy()
        child_env["PYTHONIOENCODING"] = "utf-8"
        child_env["COMPARISON_PROGRESS_FILE"] = str(progress_path)
        processor_started = time.perf_counter()
        app.logger.info(
            "Comparison script started input_bytes=%s original_name=%s",
            len(file_bytes),
            original_name,
        )
        if progress_callback:
            progress_callback(20, "Starting Comparison Engine")
        process = subprocess.Popen(
            [sys.executable, str(temp_script)],
            cwd=tmpdir,
            env=child_env,
        )
        deadline = time.monotonic() + COMPARISON_PROCESS_TIMEOUT_SECONDS
        progress_offset = 0
        last_progress = (20, "Starting Comparison Engine")
        last_heartbeat = time.monotonic()
        while process.poll() is None:
            if time.monotonic() >= deadline:
                process.kill()
                process.wait()
                raise subprocess.TimeoutExpired(
                    process.args,
                    COMPARISON_PROCESS_TIMEOUT_SECONDS,
                )

            if progress_path.exists():
                try:
                    with progress_path.open("r", encoding="utf-8") as progress_file:
                        progress_file.seek(progress_offset)
                        for progress_line in progress_file:
                            progress_data = json.loads(progress_line)
                            last_progress = (
                                int(progress_data["progress"]),
                                str(progress_data["step"]),
                            )
                            if progress_callback:
                                progress_callback(*last_progress)
                        progress_offset = progress_file.tell()
                except (OSError, KeyError, TypeError, ValueError, json.JSONDecodeError):
                    app.logger.debug("Comparison progress file was not ready")

            if progress_callback and time.monotonic() - last_heartbeat >= 15:
                progress_callback(*last_progress)
                last_heartbeat = time.monotonic()
            time.sleep(0.25)

        if process.returncode:
            raise subprocess.CalledProcessError(process.returncode, process.args)
        stats["timings_ms"]["comparison_processor"] = round(
            (time.perf_counter() - processor_started) * 1000,
            1,
        )
        app.logger.info(
            "Comparison script completed duration_ms=%s",
            stats["timings_ms"]["comparison_processor"],
        )
        output_path = Path(tmpdir) / "barc_nct_comparison.xlsx"
        if not output_path.exists():
            raise FileNotFoundError("barc_nct_comparison.xlsx was not generated.")

        output_read_started = time.perf_counter()
        xlsx_bytes = output_path.read_bytes()
        stats["timings_ms"]["output_read"] = round(
            (time.perf_counter() - output_read_started) * 1000,
            1,
        )

    stats["output_bytes"] = len(xlsx_bytes)
    stats["timings_ms"]["run_comparison_total"] = round(
        (time.perf_counter() - total_started) * 1000,
        1,
    )
    return xlsx_bytes, output_filename, stats


def parse_workbook_sheets(
    xlsx_bytes: bytes,
    file_id: str,
    channel_name: str,
    date_str: str,
    genre: str,
    timing_stats: dict = None,
):
    parse_started = time.perf_counter()
    uploaded_at = utc_now_iso()
    sheet_documents = []
    workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

    for worksheet in workbook.worksheets:
        sheet_started = time.perf_counter()
        rows_data = []
        for row in worksheet.iter_rows(values_only=True):
            rows_data.append([str(value) if value is not None else "" for value in row])

        headers = rows_data[0] if rows_data else []
        sheet_documents.append(
            {
                "file_id": file_id,
                "channel_name": channel_name,
                "genre": genre,
                "date": date_str,
                "sheet_name": worksheet.title,
                "headers": headers,
                "rows": rows_data,
                "row_count": len(rows_data),
                "col_count": len(headers),
                "uploaded_at": uploaded_at,
            }
        )
        app.logger.info(
            "Workbook sheet parsed sheet=%s rows=%s columns=%s duration_ms=%.1f",
            worksheet.title,
            len(rows_data),
            len(headers),
            (time.perf_counter() - sheet_started) * 1000,
        )

    workbook.close()
    if timing_stats is not None:
        timing_stats["workbook_parse"] = round(
            (time.perf_counter() - parse_started) * 1000,
            1,
        )
    return sheet_documents


def upload_to_db(
    xlsx_bytes: bytes,
    channel_name: str,
    date_str: str,
    original_filename: str,
    timing_stats: dict = None,
    request_id: str = "background",
):
    """Parse the processed Excel workbook and store it in MongoDB."""
    timings = timing_stats if timing_stats is not None else {}
    total_started = time.perf_counter()
    started = time.perf_counter()
    processed_files, sheets, _ = get_collections()
    timings["mongo_connection_and_indexes"] = round(
        (time.perf_counter() - started) * 1000,
        1,
    )

    started = time.perf_counter()
    channel_record = ensure_channel_record(channel_name)
    timings["mongo_channel_lookup"] = round(
        (time.perf_counter() - started) * 1000,
        1,
    )
    channel_name = channel_record["channel_name"]
    genre = channel_record["genre"]
    file_id = str(uuid.uuid4())
    uploaded_at = utc_now_iso()
    sheet_documents = parse_workbook_sheets(
        xlsx_bytes,
        file_id,
        channel_name,
        date_str,
        genre,
        timing_stats=timings,
    )

    file_document = {
        "file_id": file_id,
        "channel_name": channel_name,
        "genre": genre,
        "date": date_str,
        "original_filename": original_filename,
        "xlsx_data": Binary(xlsx_bytes),
        "uploaded_at": uploaded_at,
    }

    size_started = time.perf_counter()
    file_document_bytes = len(BSON.encode(file_document))
    sheet_document_sizes = [len(BSON.encode(document)) for document in sheet_documents]
    timings["mongo_document_sizing"] = round(
        (time.perf_counter() - size_started) * 1000,
        1,
    )
    app.logger.info(
        "MongoDB document sizes request_id=%s workbook_document_bytes=%s sheet_documents_bytes=%s max_sheet_document_bytes=%s",
        request_id,
        file_document_bytes,
        sum(sheet_document_sizes),
        max(sheet_document_sizes, default=0),
    )

    lookup = {"channel_name": channel_name, "date": date_str}
    started = time.perf_counter()
    sheets.delete_many(lookup)
    timings["mongo_delete_sheets"] = round((time.perf_counter() - started) * 1000, 1)

    started = time.perf_counter()
    processed_files.replace_one(lookup, file_document, upsert=True)
    timings["mongo_store_workbook"] = round((time.perf_counter() - started) * 1000, 1)

    if sheet_documents:
        started = time.perf_counter()
        sheets.insert_many(sheet_documents)
        timings["mongo_insert_sheets"] = round((time.perf_counter() - started) * 1000, 1)
    else:
        timings["mongo_insert_sheets"] = 0.0

    timings["mongo_storage_total"] = round(
        (time.perf_counter() - total_started) * 1000,
        1,
    )
    app.logger.info(
        "MongoDB upload complete request_id=%s channel=%s date=%s sheet_count=%s timings_ms=%s",
        request_id,
        channel_name,
        date_str,
        len(sheet_documents),
        timings,
    )
    return file_id


def update_job_progress(job_id: str, progress: int, current_step: str, worker_id: str = None):
    update = {
        "progress_percentage": max(0, min(100, int(progress))),
        "current_step": str(current_step),
        "updated_at": utc_now_iso(),
        "lease_expires_at": iso_after(seconds=JOB_LEASE_SECONDS),
    }
    query = {"job_id": job_id, "status": "PROCESSING"}
    if worker_id:
        query["worker_id"] = worker_id
    get_processing_jobs_collection().update_one(query, {"$set": update})


def read_job_file(file_id) -> bytes:
    output = io.BytesIO()
    get_job_files_bucket().download_to_stream(ObjectId(str(file_id)), output)
    return output.getvalue()


def job_status_payload(job: dict) -> dict:
    progress = int(job.get("progress_percentage") or 0)
    status = job.get("status", "FAILED")
    estimated_seconds_remaining = None
    started_at = job.get("started_at")
    if status == "PROCESSING" and started_at and 0 < progress < 100:
        try:
            elapsed = (datetime.utcnow() - datetime.fromisoformat(started_at)).total_seconds()
            estimated_seconds_remaining = max(
                0,
                int((elapsed / progress) * (100 - progress)),
            )
        except (TypeError, ValueError):
            pass

    download_available = status == "COMPLETED" and bool(job.get("result_file_id"))
    retry_available = status == "FAILED" and bool(job.get("input_file_id"))
    payload = {
        "job_id": job.get("job_id"),
        "status": status,
        "progress": progress,
        "progress_percentage": progress,
        "current_step": job.get("current_step", ""),
        "error": job.get("error_message"),
        "error_message": job.get("error_message"),
        "created_at": job.get("created_at"),
        "started_at": job.get("started_at"),
        "completed_at": job.get("completed_at"),
        "estimated_seconds_remaining": estimated_seconds_remaining,
        "download_available": download_available,
        "retry_available": retry_available,
        "attempt_count": int(job.get("attempt_count") or 0),
        "original_filename": job.get("original_filename", ""),
        "output_filename": job.get("output_filename", ""),
        "channel": job.get("channel", ""),
        "date": job.get("date", ""),
        "report_file_id": job.get("report_file_id"),
    }
    if download_available:
        payload["download_url"] = f"/api/jobs/{job['job_id']}/download"
    if retry_available:
        payload["retry_url"] = f"/api/jobs/{job['job_id']}/retry"
    return payload


def create_processing_job(uploaded_file) -> dict:
    original_filename = Path(uploaded_file.filename or "upload.xlsx").name
    extension = Path(original_filename).suffix.lower()
    if extension not in {".xlsx", ".xls"}:
        raise ValueError("Only .xlsx and .xls files are supported")

    job_id = str(uuid.uuid4())
    now = utc_now_iso()
    bucket = get_job_files_bucket()
    input_file_id = bucket.upload_from_stream(
        original_filename,
        uploaded_file.stream,
        metadata={
            "job_id": job_id,
            "kind": "input",
            "created_at": now,
        },
    )

    file_metadata = get_mongo_database().job_files.files.find_one(
        {"_id": input_file_id},
        {"length": 1},
    ) or {}
    input_size = int(file_metadata.get("length") or 0)
    if input_size <= 0:
        bucket.delete(input_file_id)
        raise ValueError("Uploaded file is empty")

    job = {
        "job_id": job_id,
        "created_at": now,
        "started_at": None,
        "completed_at": None,
        "updated_at": now,
        "status": "QUEUED",
        "progress_percentage": 0,
        "current_step": "Waiting for worker",
        "error_message": None,
        "original_filename": original_filename,
        "input_file_id": input_file_id,
        "input_size_bytes": input_size,
        "input_expires_at": iso_after(days=JOB_UPLOAD_RETENTION_DAYS),
        "result_file_id": None,
        "report_file_id": None,
        "output_filename": None,
        "channel": None,
        "date": None,
        "attempt_count": 0,
        "worker_id": None,
        "lease_expires_at": None,
        "timings_ms": {},
    }
    try:
        get_processing_jobs_collection().insert_one(job)
    except Exception:
        bucket.delete(input_file_id)
        raise
    return job


def clean_num(val):
    if not val or val == "None" or val == "nan":
        return 0
    text = str(val).replace(",", "").strip()
    try:
        return int(float(text))
    except (ValueError, TypeError):
        return 0


def parse_count_cell(value) -> int:
    text = str(value or "").strip()
    if text in {"", "-", "\u2013", "\u2014", "\u00e2\u20ac\u201d"}:
        return 0
    match = re.search(r"-?\d+(?:\.\d+)?", text.replace(",", ""))
    if not match:
        return 0
    try:
        return int(float(match.group(0)))
    except (TypeError, ValueError):
        return 0


def workbook_bytes_from_document(document) -> bytes:
    return bytes(document.get("xlsx_data") or b"")


@app.route("/")
def index():
    return send_from_directory(str(BASE_DIR), "content_analyzer.html")


@app.route("/healthz")
def healthz():
    return json_response({"status": "ok"})


@app.route("/api/genres", methods=["GET"])
def get_genres():
    """Return the complete supported genre hierarchy."""
    try:
        migrate_genre_hierarchy()
        master = get_channel_master_collection()
        counts = {
            row["_id"]: row["count"]
            for row in master.aggregate(
                [
                    {"$match": {"status": {"$ne": "inactive"}}},
                    {"$group": {"_id": "$genre", "count": {"$sum": 1}}},
                ]
            )
            if row.get("_id") in SUPPORTED_GENRES
        }
        return json_response(
            [{"genre": genre, "channel_count": counts.get(genre, 0)} for genre in SUPPORTED_GENRES]
        )
    except (PyMongoError, RuntimeError) as exc:
        return database_error_response(exc)


@app.route("/api/channels", methods=["GET"])
def get_channels_by_genre():
    """Return master channels belonging to one validated genre."""
    try:
        genre = validate_genre(request.args.get("genre", ""))
        migrate_genre_hierarchy()
        master = get_channel_master_collection()
        docs = master.find(
            {"genre": genre, "status": {"$ne": "inactive"}},
            {
                "_id": 0,
                "channel_name": 1,
                "genre": 1,
                "status": 1,
                "created_at": 1,
                "updated_at": 1,
            },
        ).sort("channel_name", ASCENDING)
        return json_response(list(docs))
    except ValueError as exc:
        return api_error(str(exc), 400)
    except (PyMongoError, RuntimeError) as exc:
        return database_error_response(exc)


@app.route("/api/channel-details", methods=["GET", "PUT"])
def channel_details():
    """Get or update one channel master record."""
    try:
        migrate_genre_hierarchy()
        master = get_channel_master_collection()

        if request.method == "GET":
            name = request.args.get("channel", "")
            key = channel_key(name)
            if not key:
                return api_error("channel required", 400)
            record = master.find_one(
                {"channel_key": key},
                {
                    "_id": 0,
                    "channel_key": 0,
                },
            )
            if not record:
                return api_error("Channel not found", 404)
            return json_response(record)

        data = request.get_json(silent=True) or {}
        name = data.get("channel_name", "")
        genre = data.get("genre", "")
        status = data.get("status", "active")
        record = update_channel_assignment(name, genre, status)
        return json_response(
            {
                key: value
                for key, value in record.items()
                if key not in {"_id", "channel_key"}
            }
        )
    except ValueError as exc:
        return api_error(str(exc), 400)
    except (PyMongoError, RuntimeError) as exc:
        return database_error_response(exc)


@app.route("/api/channel-dates", methods=["GET"])
def get_channel_dates():
    """Return available report dates for one master channel."""
    channel = normalize_channel_name(request.args.get("channel", ""))
    if not channel:
        return api_error("channel required", 400)

    try:
        migrate_genre_hierarchy()
        master = get_channel_master_collection()
        record = master.find_one({"channel_key": channel_key(channel)})
        if not record or record.get("status") == "inactive":
            return api_error("Channel not found", 404)

        processed_files, _, _ = get_collections()
        dates = processed_files.distinct("date", {"channel_name": record["channel_name"]})
        return json_response(
            {
                "channel_name": record["channel_name"],
                "genre": record["genre"],
                "dates": sorted(date for date in dates if date),
            }
        )
    except (PyMongoError, RuntimeError) as exc:
        return database_error_response(exc)


@app.route("/api/channels-dates", methods=["GET"])
def get_channels_dates():
    """Return distinct channel/date combinations from MongoDB."""
    try:
        migrate_genre_hierarchy()
        processed_files, _, _ = get_collections()
        master = get_channel_master_collection()
        channel_records = {
            row["channel_key"]: row
            for row in master.find(
                {"status": {"$ne": "inactive"}},
                {"_id": 0, "channel_key": 1, "genre": 1, "status": 1},
            )
        }
        docs = processed_files.find(
            {},
            {"_id": 0, "channel_name": 1, "date": 1},
        ).sort([("channel_name", ASCENDING), ("date", ASCENDING)])
        result = []
        for doc in docs:
            name = doc.get("channel_name", "")
            record = channel_records.get(channel_key(name))
            if not record:
                continue
            result.append(
                {
                    "channel_name": name,
                    "date": doc.get("date", ""),
                    "genre": record["genre"],
                    "status": record["status"],
                }
            )
        return json_response(result)
    except (PyMongoError, RuntimeError) as exc:
        return database_error_response(exc)


@app.route("/api/dashboard", methods=["GET"])
def get_dashboard():
    """Return KPI data for the dashboard."""
    channel = request.args.get("channel", "")
    date = request.args.get("date", "")
    source = request.args.get("source", "TABSONS-BARC")
    data_type = request.args.get("data_type", "COUNT")

    if not channel or not date:
        return api_error("channel and date required", 400)

    try:
        _, sheets, _ = get_collections()
        row = sheets.find_one(
            {"channel_name": channel, "date": date, "sheet_name": "TABSONS SUMMARY"},
            {"_id": 0, "headers": 1, "rows": 1},
        )
    except (PyMongoError, RuntimeError) as exc:
        return database_error_response(exc)

    if not row:
        return api_error("No data found for this channel/date", 404)

    all_rows = row.get("rows") or []
    headers = all_rows[1] if len(all_rows) > 1 else row.get("headers", [])
    data_row = all_rows[2] if len(all_rows) > 2 else []

    def get_val(header_keyword, r=data_row, hdrs=headers):
        for i, header in enumerate(hdrs):
            if header_keyword.upper() in str(header).upper() and i < len(r):
                return r[i]
        return "0"

    result = {"source": source, "data_type": data_type}

    if source == "TABSONS":
        if data_type == "COUNT":
            result["total_line_item"] = clean_num(get_val("TABSONS LINE ITEM"))
            result["commercial"] = clean_num(get_val("TABSONS COMMERCIAL COUNT"))
            result["promo"] = clean_num(get_val("TABSONS PROMO COUNT"))
            result["promo_sponsor"] = clean_num(get_val("TABSONS PROMO SPONSOR COUNT"))
            result["program"] = clean_num(get_val("TABSONS PROGRAM COUNT"))
        else:
            result["total_line_item"] = get_val("TABSONS DURATION")
            result["commercial"] = get_val("TABSONS COMMERCIAL DURATION")
            result["promo"] = get_val("TABSONS PROMO DURATION")
            result["promo_sponsor"] = get_val("TABSONS PROMO SPONSOR COUNT DURATIO")
            result["program"] = get_val("TABSONS PROGRAM DURATION")

    elif source == "BARC XML":
        if data_type == "COUNT":
            result["total_line_item"] = clean_num(get_val("BARC LINE ITEM"))
            result["commercial"] = clean_num(get_val("BARC COMMERCIAL COUNT"))
            result["promo"] = clean_num(get_val("BARC PROMO COUNT"))
            result["promo_sponsor"] = clean_num(get_val("BARC PROMO SPONSOR COUNT"))
            result["program"] = clean_num(get_val("BARC PROGRAM COUNT"))
        else:
            result["total_line_item"] = get_val("BARC DURATION")
            result["commercial"] = get_val("BARC COMMERCIAL DURATION")
            result["promo"] = get_val("BARC PROMO DURATION")
            result["promo_sponsor"] = get_val("BARC PROMO SPONSOR COUNT DURATION")
            result["program"] = get_val("BARC PROGRAM DURATION")

    else:
        if data_type == "COUNT":
            result["tabsons_total"] = clean_num(get_val("TABSONS LINE ITEM"))
            result["tabsons_commercial"] = clean_num(get_val("TABSONS COMMERCIAL COUNT"))
            result["tabsons_promo"] = clean_num(get_val("TABSONS PROMO COUNT"))
            result["tabsons_promo_sponsor"] = clean_num(get_val("TABSONS PROMO SPONSOR COUNT"))
            result["tabsons_program"] = clean_num(get_val("TABSONS PROGRAM COUNT"))
            result["barc_total"] = clean_num(get_val("BARC LINE ITEM"))
            result["barc_commercial"] = clean_num(get_val("BARC COMMERCIAL COUNT"))
            result["barc_promo"] = clean_num(get_val("BARC PROMO COUNT"))
            result["barc_promo_sponsor"] = clean_num(get_val("BARC PROMO SPONSOR COUNT"))
            result["barc_program"] = clean_num(get_val("BARC PROGRAM COUNT"))
        else:
            result["tabsons_total"] = get_val("TABSONS DURATION")
            result["tabsons_commercial"] = get_val("TABSONS COMMERCIAL DURATION")
            result["tabsons_promo"] = get_val("TABSONS PROMO DURATION")
            result["tabsons_promo_sponsor"] = get_val("TABSONS PROMO SPONSOR COUNT DURATIO")
            result["tabsons_program"] = get_val("TABSONS PROGRAM DURATION")
            result["barc_total"] = get_val("BARC DURATION")
            result["barc_commercial"] = get_val("BARC COMMERCIAL DURATION")
            result["barc_promo"] = get_val("BARC PROMO DURATION")
            result["barc_promo_sponsor"] = get_val("BARC PROMO SPONSOR COUNT DURATION")
            result["barc_program"] = get_val("BARC PROGRAM DURATION")

    return json_response(result)


@app.route("/api/sheets", methods=["GET"])
def get_sheets():
    """Return available sheets for a channel/date."""
    channel = request.args.get("channel", "")
    date = request.args.get("date", "")

    if not channel or not date:
        return api_error("channel and date required", 400)

    try:
        _, sheets, _ = get_collections()
        docs = sheets.find(
            {"channel_name": channel, "date": date},
            {"_id": 0, "sheet_name": 1, "row_count": 1, "col_count": 1},
        ).sort([("_id", ASCENDING)])
        return json_response(
            [
                {
                    "sheet_name": doc.get("sheet_name", ""),
                    "row_count": doc.get("row_count", 0),
                    "col_count": doc.get("col_count", 0),
                }
                for doc in docs
            ]
        )
    except (PyMongoError, RuntimeError) as exc:
        return database_error_response(exc)


@app.route("/api/sheet-data", methods=["GET"])
def get_sheet_data():
    """Return full sheet data for a channel/date/sheet."""
    channel = request.args.get("channel", "")
    date = request.args.get("date", "")
    sheet = request.args.get("sheet", "")

    if not channel or not date or not sheet:
        return api_error("channel, date, and sheet required", 400)

    try:
        _, sheets, _ = get_collections()
        row = sheets.find_one(
            {"channel_name": channel, "date": date, "sheet_name": sheet},
            {
                "_id": 0,
                "file_id": 1,
                "channel_name": 1,
                "date": 1,
                "sheet_name": 1,
                "headers": 1,
                "rows": 1,
                "row_count": 1,
                "col_count": 1,
                "uploaded_at": 1,
            },
        )
    except (PyMongoError, RuntimeError) as exc:
        return database_error_response(exc)

    if not row:
        return api_error("Sheet not found", 404)

    return json_response(
        {
            "file_id": row.get("file_id", ""),
            "channel_name": row.get("channel_name", ""),
            "date": row.get("date", ""),
            "sheet_name": row.get("sheet_name", ""),
            "headers": row.get("headers", []),
            "rows": row.get("rows", []),
            "row_count": row.get("row_count", 0),
            "col_count": row.get("col_count", 0),
            "uploaded_at": row.get("uploaded_at", ""),
        }
    )


@app.route("/api/commercial-comparison", methods=["GET"])
def get_commercial_comparison():
    """Return matched and unmatched brand data from COMMERCIAL COMPARISION sheet."""
    channel = request.args.get("channel", "")
    date = request.args.get("date", "")

    if not channel or not date:
        return api_error("channel and date required", 400)

    try:
        _, sheets, brand_modifications = get_collections()
        row = sheets.find_one(
            {"channel_name": channel, "date": date, "sheet_name": "COMMERCIAL COMPARISION"},
            {"_id": 0, "rows": 1},
        )
        mods = list(
            brand_modifications.find(
                {"channel_name": channel, "date": date},
                {"_id": 0},
            ).sort([("timestamp", ASCENDING)])
        )
    except (PyMongoError, RuntimeError) as exc:
        return database_error_response(exc)

    if not row:
        return api_error("Commercial comparison data not found", 404)

    rows = row.get("rows") or []
    headers_row = None
    matched_rows = []
    unmatched_rows = []
    section = "none"

    for r in rows:
        first_cell = str(r[0]).strip() if r else ""

        if first_cell == "SOURCE" and headers_row is None:
            headers_row = r
            continue

        if "BARC COMMERCIAL vs NCT COMMERCIAL" in first_cell and "MATCHED" in first_cell:
            section = "matched"
            continue

        if "NCT COMMERCIAL BRANDS" in first_cell and "NOT MATCHED" in first_cell:
            section = "unmatched"
            continue

        if first_cell == "MATCHED" or (
            "MATCHED" in first_cell
            and "UNMATCHED" not in first_cell
            and "NCT COMMERCIAL BRANDS" not in first_cell
            and "BARC COMMERCIAL" not in first_cell
            and first_cell not in ("MATCHING COMMERCIAL TOTAL",)
        ):
            section = "matched"
            continue

        if "NOT MATCHED" in first_cell or "UNMATCHED" in first_cell:
            section = "unmatched"
            continue

        if first_cell in (
            "",
            "MATCHING COMMERCIAL TOTAL",
            "NCT UNMATCHED TOTAL",
            "GRAND TOTAL",
        ) or "COMMERCIAL COMPARISION" in first_cell:
            continue

        if headers_row and len(r) >= 5:
            row_dict = {}
            header_labels = [
                "source",
                "channel_name",
                "date",
                "barc_brand",
                "nct_brand",
                "barc_count",
                "nct_count",
                "barc_duration",
                "nct_duration",
                "nct_ps_count",
                "nct_ps_duration",
                "remarks",
            ]
            for j, label in enumerate(header_labels):
                if j < len(r):
                    row_dict[label] = str(r[j]) if r[j] is not None else ""

            if section == "matched" and first_cell in ("BARC XML", "NCT"):
                matched_rows.append(row_dict)
            elif section == "unmatched" and first_cell in ("NCT", "BARC XML"):
                unmatched_rows.append(row_dict)

    for mod in mods:
        action = mod.get("action")
        brand_name = mod.get("brand_name", "")

        if action == "remove_from_matched":
            removed = None
            for j, row_data in enumerate(matched_rows):
                if row_data.get("barc_brand", "").strip() == brand_name.strip():
                    removed = matched_rows.pop(j)
                    break

            if not removed:
                for j, row_data in enumerate(matched_rows):
                    if row_data.get("nct_brand", "").strip() == brand_name.strip():
                        removed = matched_rows.pop(j)
                        break

            if removed:
                removed["remarks"] = "REMOVED FROM MATCHED"
                unmatched_rows.append(removed)

        elif action == "merge_to_matched":
            target_barc = mod.get("target_barc_brand") or ""
            merged = None
            for j, row_data in enumerate(unmatched_rows):
                if row_data.get("nct_brand", "").strip() == brand_name.strip():
                    merged = unmatched_rows.pop(j)
                    break

            if merged and target_barc:
                for row_data in matched_rows:
                    if row_data.get("barc_brand", "").strip() == target_barc.strip():
                        existing_count = parse_count_cell(row_data.get("nct_count", "0"))
                        merge_count = parse_count_cell(merged.get("nct_count", "0"))
                        row_data["nct_count"] = str(existing_count + merge_count)
                        row_data["remarks"] = "MATCHED (MERGED)"
                        break

    return json_response(
        {
            "matched": matched_rows,
            "unmatched": unmatched_rows,
            "headers": headers_row or [],
        }
    )


@app.route("/api/commercial/move-brand", methods=["POST"])
def move_brand():
    """Move a brand between matched and unmatched tables."""
    data = request.get_json(silent=True) or {}
    channel = data.get("channel", "")
    date = data.get("date", "")
    action = data.get("action", "")
    brand_name = data.get("brand_name", "")
    target_barc_brand = data.get("target_barc_brand", "")

    if not channel or not date or not action or not brand_name:
        return api_error("Missing required fields", 400)

    try:
        _, _, brand_modifications = get_collections()
        brand_modifications.insert_one(
            {
                "channel_name": channel,
                "date": date,
                "action": action,
                "brand_name": brand_name,
                "target_barc_brand": target_barc_brand,
                "timestamp": utc_now_iso(),
            }
        )
    except (PyMongoError, RuntimeError) as exc:
        return database_error_response(exc)

    return json_response({"success": True, "message": f"Brand '{brand_name}' {action} successfully"})


@app.route("/api/commercial/undo-modifications", methods=["POST"])
def undo_modifications():
    """Clear all brand modifications for a channel/date."""
    data = request.get_json(silent=True) or {}
    channel = data.get("channel", "")
    date = data.get("date", "")

    try:
        _, _, brand_modifications = get_collections()
        brand_modifications.delete_many({"channel_name": channel, "date": date})
    except (PyMongoError, RuntimeError) as exc:
        return database_error_response(exc)

    return json_response({"success": True})


@app.route("/api/download/preprocessed", methods=["GET"])
def download_preprocessed():
    """Download the original processed Excel file."""
    channel = request.args.get("channel", "")
    date = request.args.get("date", "")

    try:
        processed_files, _, _ = get_collections()
        row = processed_files.find_one(
            {"channel_name": channel, "date": date},
            {"_id": 0, "xlsx_data": 1, "original_filename": 1},
        )
    except (PyMongoError, RuntimeError) as exc:
        return database_error_response(exc)

    if not row:
        return api_error("File not found", 404)

    return send_file(
        io.BytesIO(workbook_bytes_from_document(row)),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=row.get("original_filename") or "report.xlsx",
    )


@app.route("/api/download/updated", methods=["GET"])
def download_updated():
    """Download the updated Excel report with brand modifications applied."""
    channel = request.args.get("channel", "")
    date = request.args.get("date", "")

    try:
        processed_files, _, brand_modifications = get_collections()
        row = processed_files.find_one(
            {"channel_name": channel, "date": date},
            {"_id": 0, "xlsx_data": 1, "original_filename": 1},
        )
        mods = list(
            brand_modifications.find(
                {"channel_name": channel, "date": date},
                {"_id": 0},
            ).sort([("timestamp", ASCENDING)])
        )
    except (PyMongoError, RuntimeError) as exc:
        return database_error_response(exc)

    if not row:
        return api_error("File not found", 404)

    xlsx_data = workbook_bytes_from_document(row)

    if mods:
        import openpyxl

        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_data))

        if "COMMERCIAL COMPARISION" in workbook.sheetnames:
            worksheet = workbook["COMMERCIAL COMPARISION"]
            last_row = worksheet.max_row + 2
            worksheet.cell(row=last_row, column=1, value="MODIFICATIONS APPLIED:")
            for i, mod in enumerate(mods):
                worksheet.cell(
                    row=last_row + i + 1,
                    column=1,
                    value=f"{mod.get('action')}: {mod.get('brand_name')}",
                )
                target_barc_brand = mod.get("target_barc_brand")
                if target_barc_brand:
                    worksheet.cell(
                        row=last_row + i + 1,
                        column=2,
                        value=f"\u2192 {target_barc_brand}",
                    )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        xlsx_data = output.read()

    filename = row.get("original_filename") or "report.xlsx"
    if mods:
        base, ext = os.path.splitext(filename)
        filename = f"{base}_UPDATED{ext}"

    return send_file(
        io.BytesIO(xlsx_data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/analyze", methods=["POST"])
def analyze():
    """Process uploaded file(s), store generated workbook data, and return downloads."""
    files = request.files.getlist("files")
    if not files:
        return api_error("No files provided", 400)

    results = []
    errors = []

    for uploaded_file in files:
        try:
            file_bytes = uploaded_file.read()
            xlsx_bytes, output_filename, stats = run_comparison(file_bytes, uploaded_file.filename)

            channel = stats.get("channel", "UNKNOWN")
            date = stats.get("date", "00/00/0000")
            try:
                file_id = upload_to_db(xlsx_bytes, channel, date, output_filename)
                stats["uploaded_to_db"] = True
                stats["file_id"] = file_id
            except Exception as db_err:
                stats["uploaded_to_db"] = False
                stats["db_error"] = str(db_err)
                app.logger.exception("MongoDB upload failed: %s", db_err)

            results.append({"fname": output_filename, "data": xlsx_bytes, "stats": stats})
        except Exception as exc:
            errors.append({"file": uploaded_file.filename, "error": str(exc)})

    if len(files) == 1 and len(results) == 1:
        result = results[0]
        return send_file(
            io.BytesIO(result["data"]),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=result["fname"],
        )

    if results:
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for result in results:
                zip_file.writestr(result["fname"], result["data"])
        zip_buf.seek(0)
        zip_name = f"barc_nct_batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        return send_file(
            zip_buf,
            mimetype="application/zip",
            as_attachment=True,
            download_name=zip_name,
        )

    return api_error("All files failed", 500, details=errors)


@app.route("/api/compare", methods=["POST"])
def compare_report():
    """Persist an upload as a durable processing job and return immediately."""
    uploaded_file = request.files.get("file")
    if not uploaded_file:
        return api_error("No file provided", 400)

    try:
        job = create_processing_job(uploaded_file)
        payload = job_status_payload(job)
        payload.update(
            {
                "success": True,
                "status_url": f"/api/job-status/{job['job_id']}",
            }
        )
        return json_response(payload, 202)
    except ValueError as exc:
        return api_error(str(exc), 400)
    except (PyMongoError, RuntimeError) as exc:
        return database_error_response(exc)


@app.route("/api/job-status/<job_id>", methods=["GET"])
@app.route("/api/jobs/<job_id>", methods=["GET"])
def get_job_status(job_id):
    try:
        job = get_processing_jobs_collection().find_one({"job_id": job_id})
    except (PyMongoError, RuntimeError) as exc:
        return database_error_response(exc)
    if not job:
        return api_error("Job not found", 404)
    return json_response(job_status_payload(job))


@app.route("/api/jobs", methods=["GET"])
def get_job_history():
    try:
        limit = min(max(int(request.args.get("limit", "20")), 1), 100)
        jobs = get_processing_jobs_collection().find(
            {},
            {
                "_id": 0,
                "worker_id": 0,
                "lease_expires_at": 0,
            },
        ).sort("created_at", -1).limit(limit)
        return json_response([job_status_payload(job) for job in jobs])
    except ValueError:
        return api_error("limit must be an integer", 400)
    except (PyMongoError, RuntimeError) as exc:
        return database_error_response(exc)


@app.route("/api/jobs/<job_id>/retry", methods=["POST"])
def retry_job(job_id):
    try:
        now = utc_now_iso()
        job = get_processing_jobs_collection().find_one_and_update(
            {
                "job_id": job_id,
                "status": "FAILED",
                "input_file_id": {"$ne": None},
            },
            {
                "$set": {
                    "status": "QUEUED",
                    "progress_percentage": 0,
                    "current_step": "Waiting for worker",
                    "error_message": None,
                    "started_at": None,
                    "completed_at": None,
                    "updated_at": now,
                    "worker_id": None,
                    "lease_expires_at": None,
                },
                "$unset": {
                    "result_file_id": "",
                    "result_size_bytes": "",
                    "report_file_id": "",
                    "output_filename": "",
                    "channel": "",
                    "date": "",
                },
            },
            return_document=ReturnDocument.AFTER,
        )
    except (PyMongoError, RuntimeError) as exc:
        return database_error_response(exc)

    if not job:
        existing = get_processing_jobs_collection().find_one({"job_id": job_id})
        if not existing:
            return api_error("Job not found", 404)
        return api_error("Job cannot be retried", 409)
    return json_response(job_status_payload(job), 202)


@app.route("/api/jobs/<job_id>/download", methods=["GET"])
def download_job_result(job_id):
    try:
        job = get_processing_jobs_collection().find_one(
            {"job_id": job_id},
            {
                "status": 1,
                "result_file_id": 1,
                "output_filename": 1,
            },
        )
        if not job:
            return api_error("Job not found", 404)
        if job.get("status") != "COMPLETED" or not job.get("result_file_id"):
            return api_error("Download is not available", 409)
        xlsx_data = read_job_file(job["result_file_id"])
    except NoFile:
        return api_error("Generated report is no longer available", 410)
    except (PyMongoError, RuntimeError) as exc:
        return database_error_response(exc)

    return send_file(
        io.BytesIO(xlsx_data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=job.get("output_filename") or "comparison_result.xlsx",
    )


@app.route("/api/template", methods=["GET"])
def download_template():
    """Download the brand comparison template file bundled with the project."""
    if not TEMPLATE_FILE.exists():
        return api_error("Template file not found", 404)

    return send_file(
        TEMPLATE_FILE,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="brand_comparison_template.xlsx",
    )


def install_route_guards() -> None:
    """Wrap every Flask view in a route-level try/except without changing routes."""
    for endpoint, view_func in list(app.view_functions.items()):
        if endpoint == "static" or getattr(view_func, "_route_guarded", False):
            continue
        app.view_functions[endpoint] = route_guard(view_func)
    app.logger.debug("Route guards installed for Flask views")


def run_startup_diagnostics() -> None:
    """Log production startup details and validate MongoDB when configured."""
    app.logger.info(
        "Startup diagnostics flask_env=%s db_name=%s mongo_uri_set=%s template_exists=%s processor_exists=%s",
        os.getenv("FLASK_ENV", ""),
        DB_NAME,
        bool(MONGO_URI),
        TEMPLATE_FILE.exists(),
        COMPARISON_SCRIPT.exists(),
    )

    if os.getenv("MONGO_DB_NAME") and not os.getenv("DB_NAME"):
        app.logger.warning("MONGO_DB_NAME is deprecated. Set DB_NAME on Render and locally.")

    if CONFIGURED_DEFAULT_CHANNEL_GENRE not in SUPPORTED_GENRES:
        app.logger.warning(
            "Invalid DEFAULT_CHANNEL_GENRE=%s; using NEWS",
            CONFIGURED_DEFAULT_CHANNEL_GENRE,
        )

    if not MONGO_URI:
        app.logger.warning("MONGO_URI is not set. Database routes require MongoDB Atlas.")
        return

    try:
        ensure_mongo_indexes()
        migrate_genre_hierarchy()
    except Exception as exc:
        app.logger.warning("MongoDB startup validation/index initialization deferred: %s", exc)


install_route_guards()
run_startup_diagnostics()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_ENV", "").lower() == "development"
    app.run(host="0.0.0.0", port=port, debug=debug)
